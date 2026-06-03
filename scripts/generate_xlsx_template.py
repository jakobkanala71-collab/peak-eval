"""Generate a single peak_v2_corpus_template.xlsx with multiple tabs.

Optional — only runs when openpyxl is installed. The .xlsx is a one-shot
import into Google Sheets ('File → Import → Upload'), creating a Sheet with
all reference tabs in one go.

If openpyxl isn't available, prints a clear message and exits 0 — peak_v2
runs fine without this file (the CSV tabs in templates/sheet_tabs/ work too).

Run:  .venv/bin/python scripts/generate_xlsx_template.py
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
TABS_DIR = ROOT / "templates" / "sheet_tabs"
GOLDEN_CSV = ROOT / "corpus" / "golden_cases.csv"
AGENT_INPUT_CSV = ROOT / "templates" / "agent_input_template.csv"
OUTPUT = ROOT / "templates" / "peak_v2_corpus_template.xlsx"


def main() -> int:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError:
        print(
            "openpyxl not installed. Either:\n"
            "  1. pip install openpyxl  (then re-run this script)\n"
            "  2. Skip the .xlsx — import the CSVs from templates/sheet_tabs/ as\n"
            "     separate tabs in your Google Sheet ('File → Import → Upload').\n"
            "Both paths produce the same Sheet structure."
        )
        return 0

    if not TABS_DIR.exists() or not GOLDEN_CSV.exists():
        print(
            "Reference CSVs missing. Run these first:\n"
            "  python scripts/generate_starter_corpus.py\n"
            "  python scripts/generate_reference_csvs.py"
        )
        return 1

    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D5CF6")
    wrap = Alignment(wrap_text=True, vertical="top")

    def add_sheet_from_csv(title: str, csv_path: Path) -> None:
        ws = wb.create_sheet(title)
        with csv_path.open(newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            for r_idx, row in enumerate(reader, start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                    else:
                        cell.alignment = wrap
        # Column widths
        for c_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 28
        ws.freeze_panes = "A2"

    add_sheet_from_csv("instructions", TABS_DIR / "instructions.csv")
    add_sheet_from_csv("categories", TABS_DIR / "categories.csv")
    add_sheet_from_csv("tiers", TABS_DIR / "tiers.csv")
    add_sheet_from_csv("golden_cases", GOLDEN_CSV)
    add_sheet_from_csv("agent_input", AGENT_INPUT_CSV)

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
